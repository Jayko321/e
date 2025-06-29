from datetime import date, datetime
from openpyxl import load_workbook
from model import Student, db 
from pony.orm import *

def populate_db_from_excel() -> str:
    # Load an existing workbook
    wb = load_workbook(r"C:\Users\Пользователь\Downloads\общий_список_3_факультета_абитуриентов_25_06_25 (3).xlsx")
    # Get the active sheet
    sheet = wb.active
    if sheet is None:
        return "Error no active sheet is available"

    # Read data from cells
    #print(sheet.cell(row=9, column=3).value)
    start_row = 8
    end_row = 130
    start_col = 3
    end_col = 45
    
    for row in sheet.iter_rows(min_col=start_col, max_col=end_col, min_row=start_row,max_row=end_row):
        student = Student()
        for cell in row:
            idx = cell.col_idx # type: ignore
            if cell.value == None:
                continue 
            if idx == 3:
                student.MilitaryDistrict = cell.value 
            if idx == 4:
                student.MilitaryUnitName = cell.value 
            if idx == 5:
                student.MilitaryUnitNumber = cell.value 
            if idx == 7:
                student.IsConscript = cell.value == "по призыву"
            if idx == 8:
                student.PersonalNumber = cell.value
            if idx == 9:
                student.Rank = cell.value
            if idx == 11:
                student.Surname = cell.value
            if idx == 12:
                student.Name = cell.value
            if idx == 13:
                student.Patronimic = cell.value
            if idx == 14:
                student.BirthDate = cell.value
            if idx == 15:
                student.SNILS = str(cell.value)
            if idx == 24:
                student.IsVeteran = cell.value == "ВБД" 
            if idx == 25:
                student.Education = cell.value
            if idx == 26:
                student.EducationalInstitution = cell.value
            if idx == 28:
                student.PhoneNumber = str(cell.value)
            if idx == 30:
                student.ArrivalDate = cell.value
            if idx == 36:
                student.Faculty = cell.value
            if idx == 37:
                student.Post = cell.value
            if idx == 38:
                student.BirthPlace = cell.value
            if idx == 39:
                student.StateRewards = cell.value
            if idx == 40:
                student.DepartmentRewards = cell.value
            if idx == 41:
                student.IsMarried = cell.value == "женат"
            if idx == 42:
                student.IsSMOMember = cell.value == "да"
            if idx == 45:
                student.EducationEndDate = datetime(cell.value, 1, 1) # type: ignore

            print(idx, end=" ")
            print(cell.value, end= " ")

    commit()
    return ""